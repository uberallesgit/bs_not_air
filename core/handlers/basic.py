from aiogram import Bot
from aiogram.types import Message, FSInputFile
import openpyxl
import json
import os
from RDB import RDB

AKASHA = "6505383049:AAHcit-EyccZVa0hvXXPKVhBOSzhgEDeNsw"
JARVIS_TOKEN = '6357305111:AAHzb68csA1ojiDn620m7FFvDXcTP9tYu_s'
# CURRENT_BOT = AKASHA
CURRENT_BOT = JARVIS_TOKEN
bot = Bot(token=CURRENT_BOT,parse_mode="HTML")
cwd = os.getcwd()

async def get_bs_not_air(message: Message,bot: Bot):
    crimean_ne_counter = 0
    crimean_ne_list = []
    file_PATH = f'{cwd}/downloads/temp.xlsx'
    await message.reply(f"Получен файл bs_not_air...")
    file = await bot.get_file(message.document.file_id)
    await bot.download_file(file.file_path,file_PATH)
    await message.answer("Файл на диске")
    book = openpyxl.load_workbook(filename=file_PATH)
    sheet = book["OM"]
    for i in range(2, sheet.max_row):
        if sheet[f'f{i}'].value.startswith("KR"):
            if sheet[f'i{i}'].value == "1":
                bs_name = sheet[f'f{i}'].value.replace("KR","CR")
                a = []
                crimean_ne_counter += 1
                a.append(sheet[f'f{i}'].value.strip())
                a.append(RDB[bs_name]["address"])
                a.append(sheet[f'd{i}'].value)
                crimean_ne_list.append(a)
        elif sheet[f'f{i}'].value.startswith("CR") or sheet[f'f{i}'].value.startswith("SE"):
            if sheet[f'i{i}'].value == "1":
                a = []
                crimean_ne_counter += 1
                a.append(sheet[f'f{i}'].value.strip())
                a.append(RDB[sheet[f'f{i}'].value]["address"])
                a.append(sheet[f'd{i}'].value)
                crimean_ne_list.append(a)
    book.close()
    temp_file_path = f"{cwd}/downloads/temp.json"
    with open(temp_file_path, "w") as file:
        json.dump(crimean_ne_list,file,indent=4,ensure_ascii=False)
    print(crimean_ne_list)
    document = FSInputFile(path=temp_file_path)
    await bot.send_document(message.chat.id,document=document,caption="БС не в работе")

async def alm_rep(message: Message, bot: Bot):
    cell_alarms = [
        "UMTS Cell Unavailable",
        "Cell Unavailable",
        "GSM Local Cell Unusable",
        "Cell Blocked",
         # "Local Cell Capability Decline",
        # "GSM Cell out of Service"
    ]
    string = ""
    crimean_ne_counter = 0
    crimean_cell_ne_list = []

    file_PATH = f'{cwd}/downloads/alm_rep_temp.xlsx'
    await message.reply(f"Получен файл alm_rep...")
    file = await bot.get_file(message.document.file_id)
    await bot.download_file(file.file_path, file_PATH)
    await message.answer("Файл на диске")
    book = openpyxl.load_workbook(filename=file_PATH)
    sheet = book["Sheet1"]
    for i in range(2, sheet.max_row):
        if sheet[f'b{i}'].value.startswith("KR"):
            if sheet[f'c{i}'].value in cell_alarms:
                bs_name = sheet[f'b{i}'].value.replace("KR", "CR")
                a = []
                crimean_ne_counter += 1
                a.append(sheet[f'b{i}'].value.strip())
                a.append(RDB[bs_name]["address"])
                # a.append(sheet[f'c{i}'].value)
                a.append(sheet[f'c{i}'].value.replace('Cell Unavailable', "Не работают отдельные сектора LTE").replace(
                    'GSM Local Cell Unusable', "Не работают отделные сектора GSM").replace("Cell Blocked","Отдельные сектора заблокированы").replace("Local Cell Unusable","Не работают отдельные сектора UMTS"))
                a.append(sheet[f'a{i}'].value)
                crimean_cell_ne_list.append(a)

        elif sheet[f'b{i}'].value.startswith("CR") or sheet[f'b{i}'].value.startswith("SE"):
            if sheet[f'c{i}'].value in cell_alarms:
                a = []
                crimean_ne_counter += 1
                a.append(sheet[f'b{i}'].value.strip())
                a.append(RDB[sheet[f'b{i}'].value]["address"])
                a.append(sheet[f'c{i}'].value.replace('Cell Unavailable', "Не работают отдельные сектора LTE").replace(
                    'GSM Local Cell Unusable', "Не работают отделные сектора GSM").replace("Cell Blocked","Отдельные сектора заблокированы").replace("Local Cell Unusable","Не работают отдельные сектора UMTS"))
                a.append(sheet[f'a{i}'].value)
                crimean_cell_ne_list.append(a)
    original_crimean_ne_list = []
    cell_count = 1
    for alarm in crimean_cell_ne_list:
        if alarm not in original_crimean_ne_list:
            original_crimean_ne_list.append(alarm)
    #         string = f"{(str(alarm).lstrip('[').rstrip(']'))}\n\n"+string.lstrip("[").rstrip("]")
    # print(original_crimean_ne_list)
    # if len(string) > 4095:
    #     for x in range(0, len(string), 4095):
    #         await bot.send_message(message.chat.id, text=string[x:x + 4095])
    # else:
    #     await bot.send_message(message.chat.id, text=string)
    book.close()
    temp_file_path = f"{cwd}/downloads/alm_rep_temp.json"
    with open(temp_file_path, "w") as file:
        json.dump(original_crimean_ne_list, file, indent=4, ensure_ascii=False)
    document = FSInputFile(path=temp_file_path)
    await bot.send_document(message.chat.id, document=document, caption="Сектора не в работе")